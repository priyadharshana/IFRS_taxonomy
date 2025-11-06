# This script performs ETL operations on the IFRS taxonomy data.
# The logic is the same as ifrs_taxonomy/notebooks/ifrs_taxonomy_etl.ipynb plus
# 1. Run Log directly built into your ETL pipeline
# 2. Automated QA checks with fail-fast or force override option
# 3. Saving outputs as CSV/JSON for sharing and Pickle for local use

# === STANDARD IMPORTS ===
import re
import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import socket
from pathlib import Path
import sys
import argparse
import pickle
import logging
import time
import yaml

# Add project root to the Python path
# This allows imports from project imports
project_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(project_root))

# === PROJECT IMPORTS ===
import config.etl_config as etl_config


# === HELPER FUNCTIONS ===
def get_project_root() -> Path:
    """Returns the absolute path to the project root directory."""
    return Path(__file__).resolve().parent.parent.parent


def load_config(project_root: Path, logger: logging.Logger) -> dict:
    """Loads the main pipeline configuration file."""
    config_path = project_root / "config/etl_ifrs_taxonomy_pipeline_config.yaml"
    with open(config_path) as f:
        config = yaml.safe_load(f)
    logger.info(f"Config loaded from {config_path}")
    return config


def get_paths(config: dict, project_root: Path) -> dict:
    """Constructs a dictionary of data paths."""
    paths = {}
    for key in config.get("paths", {}):
        path_str = config["paths"].get(key)
        if path_str is None:
            raise KeyError(f"Key '{key}' not found in configuration.")
        paths[key] = project_root / path_str
    return paths


def load_data(paths: dict, logger: logging.Logger) -> pd.DataFrame:
    """Loads data from the specified input path."""
    input_path = paths.get("input_data")
    if input_path is None or not input_path.exists():
        raise FileNotFoundError(f"Input data file not found at {input_path}")
    df = pd.read_excel(input_path, sheet_name="Taxonomy ITI", engine="openpyxl")
    logger.info(f"Data loaded from {input_path} with {len(df)} rows.")
    return df


# === Main ETL Pipeline ===
def main():
    """Main function to run the ETL pipeline."""
    # Get project root
    project_root = get_project_root()

    # Setup a temporary logger for initial setup
    temp_logger = etl_config.setup_logger(
        "etl_setup", project_root / "logs/etl_setup.log"
    )
    # Load configuration
    config = load_config(project_root, temp_logger)
    # Get paths
    paths = get_paths(config, project_root)

    # Setup the main application logger using config paths
    log_path = paths["logs_dir"] / "ifrs_taxonomy_etl.log"
    logger = etl_config.setup_logger("ifrs_taxonomy_etl", log_path)

    logger.info("ETL process started.")

    # Record start time
    start_time = time.perf_counter()

    # File and sheet details
    source_file_config = config.get("source_file", {})
    filename = source_file_config.get("filename")
    sheet_name = source_file_config.get("sheet_name")

    if not filename:
        logger.error("Filename not found in configuration under 'source_file'.")
        raise ValueError("Missing 'filename' in configuration.")

    file_path = paths["input_dir"] / filename

    # === RUN METADATA ===
    run_id = datetime.now().strftime("%Y%m%d-%H%M")
    run_datetime = datetime.now().strftime("%Y-%m-%d %H:%M")
    operator = config.get("operator", "Automated ETL Pipeline")
    environment = socket.gethostname()

    # === SECTION 1: EXTRACT ===
    extract_start = time.perf_counter()

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    headers = [
        cell.value
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    ]

    preferred_label_idx = headers.index("Preferred label")
    concept_name_idx = (
        headers.index("Concept name") if "Concept name" in headers else None
    )
    type_idx = headers.index("Type") if "Type" in headers else None

    # Optional metadata columns we want to capture
    optional_cols = config.get("optional_cols", [])

    # Find indexes for all optional columns if present
    col_indexes = {}
    for col in optional_cols:
        col_indexes[col] = headers.index(col) if col in headers else None

    # Extract rows with indent + governance row number
    rows = []
    for excel_row_num, r in enumerate(
        ws.iter_rows(min_row=2, values_only=False), start=2
    ):
        label_cell = r[preferred_label_idx]
        row_data = {
            "excel_row": excel_row_num,
            "Concept name": (
                r[concept_name_idx].value if concept_name_idx is not None else None
            ),
            "Preferred label": label_cell.value,
            "indent": int(label_cell.alignment.indent or 0),
            "Type": r[type_idx].value if type_idx is not None else None,
        }
        # Add optional columns if present
        for col, idx in col_indexes.items():
            row_data[col] = r[idx].value if idx is not None else None
        rows.append(row_data)

    df_original = pd.DataFrame(rows)
    row_count_raw = len(df_original)
    extract_end = time.perf_counter()
    logger.info(
        f"Data extracted: {row_count_raw} rows in {extract_end - extract_start:.2f} seconds."
    )

    # === SECTION 2: GROUP HEADERS ===
    grouping_start = time.perf_counter()

    # Pattern to identify group headers. TODO: Load from config in v2 if needed.
    header_pattern = re.compile(r"^\[(\d{6})\]\s*(.*)")

    df_grouped = df_original.copy()
    df_grouped[["group_code", "group_name"]] = df_grouped["Concept name"].str.extract(
        header_pattern
    )
    df_grouped["group_code"] = df_grouped["group_code"].ffill()
    df_grouped["group_name"] = df_grouped["group_name"].ffill()
    row_count_grouped = len(df_grouped)

    # === SECTION 2: CLEAN LABELS & ABSTRACT TYPE ===
    df_cleaned = df_grouped.copy()
    df_cleaned["label_clean"] = df_cleaned["Preferred label"].astype(str).str.strip()
    is_abstract = df_cleaned["Type"].isna() & df_cleaned["label_clean"].str.contains(
        "[abstract]", regex=False
    )
    df_cleaned.loc[is_abstract, "Type"] = "abstract"
    df_cleaned = df_cleaned[df_cleaned["Type"].notna()].copy()
    row_count_cleaned = len(df_cleaned)

    grouping_end = time.perf_counter()
    logger.info(
        f"Data grouped and cleaned: {row_count_cleaned} rows in "
        f"{grouping_end - grouping_start:.2f} seconds."
    )

    # === SECTION 3: BUILD HIERARCHY ===
    hierarchy_start = time.perf_counter()

    def build_hierarchy(df, max_levels=5):
        parent_stack = []
        parent_concepts = []
        level_columns = {f"Level_{i + 1}": [] for i in range(max_levels)}
        for _, row in df.iterrows():
            level = row["indent"]
            concept_name = row["Concept name"]
            label = row["label_clean"]
            while parent_stack and parent_stack[-1][0] >= level:
                parent_stack.pop()
            parent_concept = parent_stack[-1][1] if parent_stack else None
            parent_concepts.append(parent_concept)
            for i in range(max_levels):
                if i < len(parent_stack):
                    level_columns[f"Level_{i + 1}"].append(parent_stack[i][2])
                else:
                    level_columns[f"Level_{i + 1}"].append(None)
            parent_stack.append((level, concept_name, label))
        df["parent_concept"] = parent_concepts
        for col, values in level_columns.items():
            df[col] = values
        return df

    # Get max hierarchy levels from config, with a fallback.
    # The dynamic calculation is still useful for validation or as a default.
    max_levels_from_config = config.get("max_hierarchy_levels", 5)  # Default to 5
    max_indent_from_data = df_cleaned["indent"].max() + 1

    if max_indent_from_data > max_levels_from_config:
        logger.warning(
            f"Data contains hierarchy depth ({max_indent_from_data}) "
            f"greater than configured max ({max_levels_from_config}). "
            f"Using config value."
        )
    max_levels = max_levels_from_config

    df_hierarchy = build_hierarchy(df_cleaned, max_levels=max_levels)

    # Remove Level_ columns since we have parent_concept
    level_cols = [col for col in df_hierarchy.columns if col.startswith("Level_")]
    df_hierarchy.drop(columns=level_cols, inplace=True)

    # Arrange columns: Taxonomy info first, then ELT data
    df_hierarchy_cols_order = config.get("df_hierarchy_cols_order", [])
    df_hierarchy = df_hierarchy[df_hierarchy_cols_order]

    # Count rows in the hierarchy DataFrame after final column arrangement for QA checks
    row_count_hierarchy = len(df_hierarchy)

    hierarchy_end = time.perf_counter()
    logger.info(
        f"Hierarchy built: {row_count_hierarchy} rows in "
        f"{hierarchy_end - hierarchy_start:.2f} seconds."
    )

    # === SECTION 4: FULL PATH & MATERIALIZED PATH TABLE ===
    fullpath_start = time.perf_counter()

    def df_to_tree(df):
        root = []
        stack = [(-1, root, "")]
        for _, row in df.iterrows():
            base_path = row["group_name"] or row["group_code"] or ""
            if not stack[-1][2] or stack[-1][2].startswith(base_path):
                current_path = base_path + (
                    " > " + row["label_clean"] if row["label_clean"] else ""
                )
            else:
                current_path = stack[-1][2] + " > " + row["label_clean"]
            # v2: Include all fields in the node
            node = {
                "excel_row": row["excel_row"],
                "concept_name": row["Concept name"],
                "preferred_label": row["Preferred label"],
                "label": row["label_clean"],
                "group_code": row["group_code"],
                "group_name": row["group_name"],
                "type": row["Type"],
                "full_path": current_path,
                "standard_label": row.get("Standard label"),
                "documentation_label": row.get("Documentation label"),
                "guidance_label": row.get("Guidance label"),
                "references": row.get("References"),
                "reference_links": row.get("Reference Links"),
                "children": [],
            }
            level = row["indent"]
            while stack and stack[-1][0] >= level:
                stack.pop()
            stack[-1][1].append(node)
            stack.append((level, node["children"], current_path))
        return root

    taxonomy_tree: list[dict] = df_to_tree(df_hierarchy)

    def flatten_tree_to_table(tree):
        flat_rows = []

        def recurse(nodes):
            for node in nodes:
                # Include all fields in the flat rows
                flat_rows.append(
                    {
                        "full_path": node["full_path"],
                        "excel_row": node["excel_row"],
                        "concept_name": node["concept_name"],
                        "preferred_label": node["preferred_label"],
                        "label": node["label"],
                        "group_code": node["group_code"],
                        "group_name": node["group_name"],
                        "type": node["type"],
                        "standard_label": node.get("standard_label"),
                        "documentation_label": node.get("documentation_label"),
                        "guidance_label": node.get("guidance_label"),
                        "references": node.get("references"),
                        "reference_links": node.get("reference_links"),
                    }
                )
                recurse(node["children"])

        recurse(tree)
        return pd.DataFrame(flat_rows)

    df_materialized_path = flatten_tree_to_table(taxonomy_tree)
    row_count_materialized = len(df_materialized_path)

    fullpath_end = time.perf_counter()
    logger.info(
        f"Materialized path built: {row_count_materialized} rows in "
        f"{fullpath_end - fullpath_start:.2f} seconds."
    )

    # === ARGUMENT PARSER ===
    parser = argparse.ArgumentParser(description="IFRS Taxonomy ETL")
    parser.add_argument(
        "--force", action="store_true", help="Override fail-fast QA checks"
    )
    args = parser.parse_args()

    # === SECTION 5: AUTOMATED QA CHECKS ===
    qa_start = time.perf_counter()

    qa_results = {
        "all_have_group_info": df_hierarchy["group_code"].notna().all()
        and df_hierarchy["group_name"].notna().all(),
        "indent_matches_int": pd.api.types.is_integer_dtype(df_hierarchy["indent"]),
        "no_empty_labels_unless_abstract": df_hierarchy.loc[
            df_hierarchy["Type"] != "abstract", "label_clean"
        ]
        .notna()
        .all(),
        "unique_full_path": df_materialized_path["full_path"].is_unique,
    }

    # Define critical checks that must pass for fail-fast mode
    critical_checks = config.get("critical_checks", [])

    failed_critical = [chk for chk in critical_checks if not qa_results[chk]]

    qa_end = time.perf_counter()
    logger.info(f"QA checks completed in {qa_end - qa_start:.2f} seconds.")

    # === SECTION 6: BUILD RUN LOG ===
    run_log_start = time.perf_counter()
    __version__ = config.get("etl_version", "unknown")

    run_log = {
        "run_metadata": {
            "run_id": run_id,
            "run_datetime": run_datetime,
            "operator": operator,
            "source_file": Path(file_path).name,
            "sheet_name": sheet_name,
            "etl_version": __version__,
            "environment": environment,
            "force_override": args.force,  # record if force was used
        },
        "row_counts": {
            "raw_extract": row_count_raw,
            "after_group_headers": row_count_grouped,
            "after_cleaning": row_count_cleaned,
            "after_hierarchy": row_count_hierarchy,
            "materialized_path": row_count_materialized,
        },
        "qa_results": qa_results,
        "anomalies": [],
        "sign_off": {"etl_operator": None, "reviewer": None},
    }

    run_log_end = time.perf_counter()
    logger.info(f"Run log built in {run_log_end - run_log_start:.2f} seconds.")

    # === SAVE JSON RUN LOG ===
    json_start = time.perf_counter()

    json_path = paths["logs_dir"] / f"run_log_{run_id}.json"

    # Convert all QA results to native Python bool
    # Python 3.13 and later: the built-in json module does not serialize NumPy or pandas
    # types like numpy.bool_ or pd.BooleanDtype directly as JSON booleans
    if isinstance(run_log["qa_results"], dict):
        run_log["qa_results"] = {k: bool(v) for k, v in run_log["qa_results"].items()}
    else:
        raise TypeError(
            "run_log['qa_results'] must be a dict, got {}".format(
                type(run_log["qa_results"])
            )
        )

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(run_log, f, indent=2, ensure_ascii=False)

    json_end = time.perf_counter()
    logger.info(f"Run log saved as JSON in {json_end - json_start:.2f} seconds.")

    # === SAVE MARKDOWN RUN LOG ===
    run_log_start = time.perf_counter()

    md_path = paths["logs_dir"] / f"run_log_{run_id}.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# IFRS Taxonomy ETL Run Log\n\n")
        f.write(f"**Run ID:** {run_log['run_metadata']['run_id']}\n")
        f.write(f"**Run Date/Time:** {run_log['run_metadata']['run_datetime']}\n")
        f.write(f"**Operator:** {run_log['run_metadata']['operator']}\n")
        f.write(f"**Source File:** {run_log['run_metadata']['source_file']}\n")
        f.write(f"**Sheet Name:** {run_log['run_metadata']['sheet_name']}\n")
        f.write(f"**ETL Version:** {run_log['run_metadata']['etl_version']}\n")
        f.write(f"**Environment:** {run_log['run_metadata']['environment']}\n")
        f.write(f"**Force Override:** {run_log['run_metadata']['force_override']}\n\n")

        f.write("## Row Counts\n")
        for stage, count in run_log["row_counts"].items():
            f.write(f"- **{stage.replace('_', ' ').title()}:** {count}\n")
        f.write("\n")

        f.write("## QA Results\n")
        for check, passed in run_log["qa_results"].items():
            if passed:
                status = '<span style="color:green">✅ Pass</span>'
            else:
                status = '<span style="color:red">❌ Fail</span>'
            f.write(f"- {check.replace('_', ' ').title()}: {status}\n")
        f.write("\n")

        f.write("## Anomalies & Exceptions\n")
        if run_log["anomalies"]:
            for anomaly in run_log["anomalies"]:
                f.write(f"- {anomaly}\n")
        else:
            f.write("- None recorded\n")
        f.write("\n")

        f.write("## Sign-Off\n")
        f.write("| Role | Name | Date | Signature |\n")
        f.write("|------|------|------|-----------|\n")
        f.write("| ETL Operator |  |  |  |\n")
        f.write("| Reviewer / QA |  |  |  |\n")

    logger.info(f"Run logs saved as:\n- {json_path}\n- {md_path}")

    run_log_end = time.perf_counter()
    logger.info(
        f"Run log saved as Markdown in {run_log_end - run_log_start:.2f} seconds."
    )

    # === FAIL-FAST OR FORCE ===
    if failed_critical and not args.force:
        print("\n❌ CRITICAL QA CHECKS FAILED ❌")
        for chk in failed_critical:
            print(f"- {chk} failed")
        print("\nETL aborted before producing downstream outputs.")
        sys.exit(1)
    elif failed_critical and args.force:
        print("\n⚠️ Force override enabled — continuing despite failed QA checks:")
        for chk in failed_critical:
            print(f"- {chk} failed")
    else:
        print(
            "\n✅ All critical QA checks passed — proceeding with downstream outputs..."
        )

    # === SAVE OUTPUTS (only if QA passed or force override) ===
    output_dir = paths["output_dir"]
    output_dir.mkdir(exist_ok=True)

    # Save hierarchy DataFrame

    df_hierarchy_start = time.perf_counter()

    df_hierarchy.to_csv(
        output_dir / f"df_hierarchy_{run_id}.csv", index=False, encoding="utf-8"
    )
    df_hierarchy.to_pickle(output_dir / f"df_hierarchy_{run_id}.pkl")  # Local use only

    df_hierarchy_end = time.perf_counter()
    logger.info(
        f"Hierarchy DataFrame saved in {df_hierarchy_end - df_hierarchy_start:.2f} seconds."
    )

    # Save materialized path DataFrame

    df_materialized_path_start = time.perf_counter()
    df_materialized_path.to_csv(
        output_dir / f"df_materialized_path_{run_id}.csv",
        index=False,
        encoding="utf-8",
    )
    df_materialized_path.to_pickle(
        output_dir / f"df_materialized_path_{run_id}.pkl"
    )  # Local use only

    df_materialized_path_end = time.perf_counter()
    logger.info(
        f"Materialized path DataFrame saved in "
        f"{df_materialized_path_end - df_materialized_path_start:.2f} seconds."
    )

    # Save taxonomy tree JSON
    json_start = time.perf_counter()

    with open(output_dir / f"taxonomy_tree_{run_id}.json", "w", encoding="utf-8") as f:
        json.dump(taxonomy_tree, f, indent=2, ensure_ascii=False)

    # Save taxonomy tree Pickle (local use only)
    with open(output_dir / f"taxonomy_tree_{run_id}.pkl", "wb") as f:
        pickle.dump(taxonomy_tree, f)

    json_end = time.perf_counter()
    logger.info(f"Taxonomy tree saved in {json_end - json_start:.2f} seconds.")

    print("\n📦 Outputs saved (CSV/JSON for sharing, Pickle for local use).")

    end_time = time.perf_counter()
    logger.info(f"ETL process completed in {end_time - start_time:.2f} seconds.")
    # --- End Outputs ---


if __name__ == "__main__":
    main()
